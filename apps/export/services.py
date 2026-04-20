from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from apps.students.models import Student, LevelByMonth


def generate_excel_stream():
    wb = Workbook()
    ws = wb.active
    ws.title = "Коты"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    separator_fill = PatternFill("solid", fgColor="1D4ED8")  # более насыщенный синий разделитель
    separator_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin')
    thick = Side(style='thick')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    separator_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # Базовые заголовки
    base_headers = [
        "№",
        "ФИО", "Имя", "Фамилия", "Отчество", "Возраст", "Текущий уровень", "Дата увольнения",
        "Статус", "Категория", "Курс", "Направление", "Подразделение", "Личный телефон", "Telegram",
        "Телефон родителя", "ФИО родителя", "Адрес фактический", "Адрес по прописке",
        "Создан", "Кем создан", "Изменён"
    ]

    # Заголовки календаря
    months_ru = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    calendar_headers = []
    for year in range(2023, 2027):
        for month_name in months_ru:
            calendar_headers.append(f"{month_name} {year}")

    headers = base_headers + calendar_headers

    def format_fired_date(d):
        if not d:
            return "—"
        # Автоматически определяем точность по дню
        if d.day == 1:
            return f"{months_ru[d.month - 1]} {d.year}"
        return d.strftime('%d.%m.%Y')

    # Запись заголовков
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # prepare fills for level coloring
    level_fill_map = {
        'black': PatternFill('solid', fgColor='000000'),
        'yellow': PatternFill('solid', fgColor='FFFF00'),
        'red': PatternFill('solid', fgColor='FF0000'),
        'green': PatternFill('solid', fgColor='00FF00'),
        'fired': PatternFill('solid', fgColor='FFA500'),
        '': PatternFill('solid', fgColor='CCCCCC'),  # no level
        None: PatternFill('solid', fgColor='CCCCCC'),
    }

    # Данные студентов с prefetch
    queryset = Student.objects.select_related('created_by', 'updated_by').prefetch_related('level_by_month').all()

    lbm_dict = {}
    for student in queryset:
        lbm_dict[student.id] = {}
        for lbm in student.level_by_month.all():
            lbm_dict[student.id][(lbm.year, lbm.month)] = lbm

    # Порядок категорий и их человекочитаемые названия
    category_order = [
        'college',
        'alabuga_start_sng',
        'alabuga_start_rf',
        'alabuga_mulatki',
        'patriot',
    ]
    category_names = {
        'college': 'Колледжисты',
        'alabuga_mulatki': 'АС',
        'alabuga_start_sng': 'АС',
        'patriot': 'Патриоты',
        'alabuga_start_rf': 'АС',
    }

    # Порядок уровней внутри категории
    level_order = {
        'green': 0,
        'yellow': 1,
        'red': 2,
        'black': 3,
        '': 4,          # без уровня
        None: 4,
        'fired': 5,     # уволенные в конце
    }

    row_num = 2

    today = timezone.now().date()
    # "последний месяц" = предыдущий календарный месяц
    first_this_month = today.replace(day=1)
    prev_month_end = first_this_month - timezone.timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)

    for cat in category_order:
        students_in_cat = [s for s in queryset if s.category == cat]
        # сортировка по уровню и ФИО
        students_in_cat.sort(key=lambda s: (level_order.get(s.level, 99), s.full_name))

        if students_in_cat:
            # Разделитель
            separator_text = category_names.get(cat, cat.upper())
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                if col == 1:
                    cell.value = separator_text
                cell.fill = separator_fill
                # более толстая и заметная разделительная строка
                cell.font = Font(bold=True, color="FFFFFF", size=13)
                cell.alignment = center
                cell.border = separator_border
            # увеличим высоту строки-разделителя
            ws.row_dimensions[row_num].height = 25
            row_num += 1

            # Студенты
            counter = 1
            for student in students_in_cat:
                # Всегда выводим актуальную дату увольнения из основного поля студента
                fired_date = student.fired_date

                # prepare raw values and track level keys for coloring
                base_values = [
                    counter,
                    student.full_name,
                    student.first_name,
                    student.last_name,
                    student.patronymic or "—",
                    student.age or "—",
                    student.get_level_display(),
                    format_fired_date(fired_date) if fired_date else "—",
                    student.get_status_display(),
                    student.get_category_display(),
                    student.get_course_display() or "—",
                    student.get_direction_display() or student.direction or "—",
                    student.get_subdivision_display() or student.subdivision or "—",
                    student.phone_personal,
                    student.telegram or "—",
                    student.phone_parent,
                    student.fio_parent,
                    student.address_actual or "—",
                    student.address_registered or "—",
                    student.created_at.strftime("%d.%m.%Y %H:%M"),
                    student.created_by.get_full_name() if student.created_by else "—",
                    student.updated_at.strftime("%d.%m.%Y %H:%M"),
                ]
                current_level_key = student.level or ''

                # calendar entries list of (display, level_key)
                calendar_entries = []
                student_lbm = lbm_dict.get(student.id, {})
                for year in range(2023, 2027):
                    for month in range(1, 13):
                        lbm = student_lbm.get((year, month))
                        if lbm and lbm.level:
                            display = lbm.get_level_display()
                            if lbm.level == 'fired' and lbm.fired_date:
                                display += f" ({format_fired_date(lbm.fired_date)})"
                            calendar_entries.append((display, lbm.level))
                        else:
                            calendar_entries.append(("—", None))

                # write cells one by one so we can color them
                full_values = base_values + [val for val, _ in calendar_entries]
                for col_idx, value in enumerate(full_values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # color current level column (index 7: с учётом колонки №)
                    if col_idx == 7:
                        fill = level_fill_map.get(current_level_key, None)
                        if fill:
                            cell.fill = fill
                    # color calendar columns
                    elif col_idx > len(base_headers):
                        cal_idx = col_idx - len(base_headers) - 1
                        level_key = calendar_entries[cal_idx][1]
                        fill = level_fill_map.get(level_key, None)
                        if fill:
                            cell.fill = fill
                row_num += 1
                counter += 1

    # Единая аккуратная ширина колонок (одинаковая в группе)
    for col_num, _ in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_num)
        if col_num <= len(base_headers):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 16

    # Высота строк, чтобы весь текст был виден и выглядел ровно
    for r in range(2, row_num):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 22

    # Заморозка шапки
    ws.freeze_panes = "A2"

    return wb