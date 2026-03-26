from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, Reference
from django.utils import timezone
from datetime import datetime, date
from django.db.models import Q
import calendar
from apps.students.models import Student, LevelHistory
from apps.hr_calls.models import HrCall
from apps.students.models import CATEGORY_CHOICES


MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

LEVEL_RU = {
    "green": "Зелёный",
    "yellow": "Жёлтый",
    "red": "Красный",
    "black": "Чёрный",
    "fired": "Уволенный",
    "": "Без уровня",
    None: "Без уровня",
}

STATUS_RU = {
    "active": "Активные",
    "fired": "Уволенные",
    "called_hr": "Вызваны к HR",
}


def _is_full_month(d_from, d_to):
    """Проверка: диапазон ровно соответствует календарному месяцу."""
    if not d_from or not d_to:
        return False
    last_day = calendar.monthrange(d_from.year, d_from.month)[1]
    return (
        d_from.day == 1
        and d_to.year == d_from.year
        and d_to.month == d_from.month
        and d_to.day == last_day
    )


def _count_fired_by_period(d_from, d_to):
    """
    Уволенные считаются по правилам:
    - точный день: день != 1 и fired_date == день
    - полный месяц: день == 1 (месячная точность) попавшие по year/month + точные дни в месяце
    - прочие диапазоны: все попавшие в диапазон
    """
    base = Student.objects.filter(status='fired', fired_date__isnull=False)
    if not d_from or not d_to:
        return base.count()

    if d_from == d_to:
        # Точный день: считаем только тех, у кого день != 1 (точная дата)
        return base.filter(fired_date=d_from).exclude(fired_date__day=1).count()

    if _is_full_month(d_from, d_to):
        # Полный месяц: считаем всех в диапазоне (включая месячную точность с day=1)
        return base.filter(
            fired_date__year=d_from.year,
            fired_date__month=d_from.month,
        ).count()

    # Прочие диапазоны: считаем всех в диапазоне
    return base.filter(fired_date__range=(d_from, d_to)).count()


def generate_analytics_excel(download_type='full', date_from=None, date_to=None):
    """
    Генерация Excel-файла с аналитикой.
    
    Args:
        download_type: 'full' - полная аналитика, 'month' - аналитика за месяц
        date_from: начало периода (для month)
        date_to: конец периода (для month)
    """
    wb = Workbook()
    
    if download_type == 'full':
        return _generate_full_analytics_excel(wb, date_from, date_to)
    else:
        return _generate_month_analytics_excel(wb, date_from, date_to)


def _generate_full_analytics_excel(wb, date_from=None, date_to=None):
    """Генерация полной аналитики с графиками"""
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2563eb")  # синий
    subheader_fill = PatternFill("solid", fgColor="3b82f6")  # светлее синий
    title_font = Font(bold=True, size=16, color="2563eb")
    subtitle_font = Font(bold=True, size=12, color="374151")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center = Alignment(horizontal="center", vertical="center")
    
    # Лист "Analytics"
    ws_analytics = wb.active
    ws_analytics.title = "Аналитика"
    
    # Заголовок
    ws_analytics.merge_cells('A1:H1')
    title_cell = ws_analytics['A1']
    title_cell.value = f"Аналитика - {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = title_font
    title_cell.alignment = center
    
    # Подзаголовок
    ws_analytics.merge_cells('A2:H2')
    subtitle_cell = ws_analytics['A2']
    subtitle_cell.value = "Распределение котов по уровню, статусу и категории"
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center
    
    row_num = 4
    
    # 1. По уровню
    ws_analytics.merge_cells(f'A{row_num}:H{row_num}')
    level_header = ws_analytics[f'A{row_num}']
    level_header.value = "По уровню"
    level_header.font = subtitle_font
    level_header.fill = subheader_fill
    level_header.alignment = center
    
    row_num += 1
    headers = ["Уровень", "Количество", "Процент"]
    for col, header in enumerate(headers, 1):
        cell = ws_analytics.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num += 1
    
    # Собираем данные для аналитики
    active_students = Student.objects.filter(status='active').exclude(level='fired')
    
    # Распределение по уровням
    distribution_by_level = {}
    for student in active_students:
        level = student.level
        if level in distribution_by_level:
            distribution_by_level[level] += 1
        else:
            distribution_by_level[level] = 1
    
    # Распределение по статусам
    distribution_by_status = {}
    for student in Student.objects.all():
        status = student.status
        if status in distribution_by_status:
            distribution_by_status[status] += 1
        else:
            distribution_by_status[status] = 1
    
    # Распределение по категориям
    distribution_by_category = {}
    for student in Student.objects.all():
        category = student.category
        if category in distribution_by_category:
            distribution_by_category[category] += 1
        else:
            distribution_by_category[category] = 1
    
    active_total = active_students.count() or 1
    level_data_start = row_num
    
    for level, count in distribution_by_level.items():
        percentage = round(count / active_total * 100, 1) if active_total > 0 else 0.0
        
        ws_analytics.cell(row=row_num, column=1, value=LEVEL_RU.get(level, str(level))).alignment = center
        ws_analytics.cell(row=row_num, column=2, value=count).alignment = center
        ws_analytics.cell(row=row_num, column=3, value=f"{percentage}%").alignment = center
        
        # Применяем границы
        for col in range(1, 4):
            ws_analytics.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    level_data_end = row_num - 1
    
    # Добавляем пустую строку
    row_num += 1
    
    # 2. По статусу
    ws_analytics.merge_cells(f'A{row_num}:H{row_num}')
    status_header = ws_analytics[f'A{row_num}']
    status_header.value = "По статусу"
    status_header.font = subtitle_font
    status_header.fill = subheader_fill
    status_header.alignment = center
    
    row_num += 1
    for col, header in enumerate(headers, 1):
        cell = ws_analytics.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num += 1
    status_data_start = row_num
    total = Student.objects.count() or 1
    for status, count in distribution_by_status.items():
        percentage = round(count / total * 100, 1)

        ws_analytics.cell(row=row_num, column=1, value=STATUS_RU.get(status, str(status))).alignment = center
        ws_analytics.cell(row=row_num, column=2, value=count).alignment = center
        ws_analytics.cell(row=row_num, column=3, value=f"{percentage}%").alignment = center
        
        for col in range(1, 4):
            ws_analytics.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    status_data_end = row_num - 1
    
    # Добавляем пустую строку
    row_num += 1
    
    # 3. По категории
    ws_analytics.merge_cells(f'A{row_num}:H{row_num}')
    category_header = ws_analytics[f'A{row_num}']
    category_header.value = "По категории"
    category_header.font = subtitle_font
    category_header.fill = subheader_fill
    category_header.alignment = center
    
    row_num += 1
    for col, header in enumerate(headers, 1):
        cell = ws_analytics.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num += 1
    category_data_start = row_num
    for category, count in distribution_by_category.items():
        display_name = dict(CATEGORY_CHOICES).get(category, category.capitalize())
        percentage = round(count / total * 100, 1)
        
        ws_analytics.cell(row=row_num, column=1, value=display_name).alignment = center
        ws_analytics.cell(row=row_num, column=2, value=count).alignment = center
        ws_analytics.cell(row=row_num, column=3, value=f"{percentage}%").alignment = center
        
        for col in range(1, 4):
            ws_analytics.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    category_data_end = row_num - 1
    
    # Настройка ширины колонок
    ws_analytics.column_dimensions['A'].width = 25
    ws_analytics.column_dimensions['B'].width = 15
    ws_analytics.column_dimensions['C'].width = 15
    
    # Лист "Charts"
    ws_charts = wb.create_sheet(title="Диаграммы")
    
    # Диаграмма по уровням
    level_chart = PieChart()
    level_chart.title = "Распределение по уровням"
    level_chart.style = 10
    
    level_data = Reference(ws_analytics, min_col=2, min_row=level_data_start, max_row=level_data_end)
    level_labels = Reference(ws_analytics, min_col=1, min_row=level_data_start, max_row=level_data_end)
    level_chart.add_data(level_data, titles_from_data=False)
    level_chart.set_categories(level_labels)
    level_chart.height = 15
    level_chart.width = 20
    
    ws_charts.add_chart(level_chart, "A1")
    
    # Диаграмма по статусу
    status_chart = PieChart()
    status_chart.title = "Распределение по статусам"
    status_chart.style = 11
    
    status_data = Reference(ws_analytics, min_col=2, min_row=status_data_start, max_row=status_data_end)
    status_labels = Reference(ws_analytics, min_col=1, min_row=status_data_start, max_row=status_data_end)
    status_chart.add_data(status_data, titles_from_data=False)
    status_chart.set_categories(status_labels)
    status_chart.height = 15
    status_chart.width = 20
    
    ws_charts.add_chart(status_chart, "P1")
    
    # Диаграмма по категориям
    category_chart = PieChart()
    category_chart.title = "Распределение по категориям"
    category_chart.style = 12
    
    category_data = Reference(ws_analytics, min_col=2, min_row=category_data_start, max_row=category_data_end)
    category_labels = Reference(ws_analytics, min_col=1, min_row=category_data_start, max_row=category_data_end)
    category_chart.add_data(category_data, titles_from_data=False)
    category_chart.set_categories(category_labels)
    category_chart.height = 15
    category_chart.width = 20
    
    ws_charts.add_chart(category_chart, "A20")
    
    return wb


def _generate_month_analytics_excel(wb, date_from=None, date_to=None):
    """Генерация аналитики за месяц"""
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2563eb")
    title_font = Font(bold=True, size=16, color="2563eb")
    subtitle_font = Font(bold=True, size=12, color="374151")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center = Alignment(horizontal="center", vertical="center")
    
    # Лист "Analytics"
    ws = wb.active
    ws.title = "Аналитика"
    
    # Заголовок
    if date_from and date_to:
        # Для month-загрузки ожидаем период одного календарного месяца
        period_str = f"{MONTHS_RU[date_from.month - 1]} {date_from.year}"
    else:
        period_str = f"{MONTHS_RU[timezone.now().month - 1]} {timezone.now().year}"
    
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Аналитика за {period_str}"
    title_cell.font = title_font
    title_cell.alignment = center
    
    # Подзаголовок
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = "Периодная аналитика: уволенные, вызовы к HR, новые коты, смены уровней"
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center
    
    row_num = 4
    
    # Данные за период
    if date_from and date_to:
        start_dt = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_dt = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
        
        # Уволенные по правилам (day vs month+year точность)
        fired_students_period = _count_fired_by_period(date_from, date_to)
        
        # Подсчет ВСЕХ вызовов к HR (независимо от решения проблемы)
        called_hr_total = HrCall.objects.filter(
            created_at__range=(start_dt, end_dt),
            person_type='cat'
        ).count()
        
        new_students_total = Student.objects.filter(created_at__range=(start_dt, end_dt)).count()
        level_changes_total = LevelHistory.objects.filter(changed_at__range=(start_dt, end_dt)).count()
    else:
        # Текущий месяц
        now = timezone.now()
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = (start_dt + timezone.timedelta(days=32)).replace(day=1)
        current_month_from = start_dt.date()
        current_month_to = (end_dt - timezone.timedelta(days=1)).date()
        
        fired_students_period = _count_fired_by_period(current_month_from, current_month_to)
        
        called_hr_total = HrCall.objects.filter(
            created_at__range=(start_dt, end_dt),
            person_type='cat'
        ).count()
        
        new_students_total = Student.objects.filter(created_at__range=(start_dt, end_dt)).count()
        level_changes_total = LevelHistory.objects.filter(changed_at__range=(start_dt, end_dt)).count()
    
    # Таблица с данными
    headers = ["Метрика", "Значение", "Описание"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num += 1
    
    metrics = [
        ("Уволенные коты", fired_students_period, "Коты, уволенные в период"),
        ("Вызовы к HR", called_hr_total, "Все вызовы к HR в период (независимо от результата)"),
        ("Новые коты", new_students_total, "Коты, созданные в период"),
        ("Смены уровней", level_changes_total, "Все изменения уровней за период"),
    ]
    
    for metric, value, description in metrics:
        ws.cell(row=row_num, column=1, value=metric).alignment = center
        ws.cell(row=row_num, column=2, value=value).alignment = center
        ws.cell(row=row_num, column=3, value=description).alignment = center
        
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    return wb