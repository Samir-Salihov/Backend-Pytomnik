from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.hr_calls.models import HrCall


def generate_hr_calls_excel_stream():
    wb = Workbook()
    ws = wb.active
    ws.title = "Вызовы к HR"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")  # dark gray
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "№",
        "ФИО",
        "Тип (Кот/Не кот)",
        "Категория",
        "Причина",
        "Решение",
        "Дата/время посещения",
        "Проблема решена",
        "Кем создан",
        "Создано",
        "Изменено",
    ]

    # header row
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    calls = HrCall.objects.select_related("created_by", "student").all().order_by("-created_at")

    row_num = 2
    for idx, call in enumerate(calls, start=1):
        full_name = call.student.full_name if (call.person_type == "cat" and call.student) else call.full_name
        person_type_display = dict(HrCall._meta.get_field("person_type").choices).get(call.person_type, call.person_type)
        # category хранит ключ (например, 'college'); для Excel показываем человекочитаемое значение
        category_key = call.category or ""
        if not category_key and call.student:
            category_key = getattr(call.student, "category", "") or ""
        category = dict(HrCall._meta.get_field("category").choices).get(category_key, category_key)

        values = [
            idx,
            full_name,
            person_type_display,
            category,
            call.reason or "—",
            call.solution or "—",
            call.visit_datetime.strftime("%d.%m.%Y %H:%M") if call.visit_datetime else "—",
            "Да" if call.problem_resolved else "Нет",
            call.created_by.get_full_name() if call.created_by else "—",
            call.created_at.strftime("%d.%m.%Y %H:%M") if call.created_at else "—",
            call.updated_at.strftime("%d.%m.%Y %H:%M") if call.updated_at else "—",
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center if col not in (5, 6) else Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row_num].height = 38
        row_num += 1

    # Column widths: fixed so text stays readable
    widths = {
        1: 6,    # №
        2: 28,   # ФИО
        3: 14,   # тип
        4: 18,   # категория
        5: 40,   # причина
        6: 40,   # решение
        7: 20,   # визит
        8: 16,   # решено
        9: 22,   # кем создан
        10: 20,  # создано
        11: 20,  # изменено
    }
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = widths.get(col_num, 18)

    ws.freeze_panes = "A2"
    return wb

